#!/usr/bin/python
# -*- coding: utf-8 -*-

import configparser
import sys,os
import xlrd,datetime
import shutil
import openpyxl
from copy import deepcopy

#定义软件位置
exe_path =os.path.split(os.path.abspath(sys.argv[0]))[0]



#读取配置文件Config.ini
config=configparser.ConfigParser()
config.read(exe_path + "\\" + "Config.ini", encoding="utf-8-sig")



#工参文件
bi_path = config.get('main', 'basic_information')
bsc_file = config.get('main', 'input_file')
#out_path = config.get('main', 'out_path')

#ddd.encode("gb18030").decode("gb18030")
bsc_file = bsc_file.split(",")
#print(bsc_file)

def readexecel(filename):
    #file_name = filename.encoding('utf-8')
    workbook = xlrd.open_workbook(exe_path +"\\" +filename) # ='UTF-8'
    #file_name = filename.split("\\")[-1]
    print("filename:",filename)
    target_cel = {}
    cel_p ={}

    if  filename == '工参_0915.xlsx':
        sheet = workbook.sheet_by_name('退频方案')
        #print("获取表名")
        rows =sheet.col_values(0)
        print(len(rows),type(rows))
        #num_list = [n for n in range(1,10000)]
        #print(num_list)
        num =0
        for  n in range(1,len(rows)):
            #读取单元格
            #print(n)
            try:
                uu = sheet.cell(n, 0)
                # print(n,sheet.cell(n,0))
                # print(n,uu)
                # print(type(uu))
                if uu == '':
                    pass
                else:
                    # print(n)
                    num +=1
                    #print(num)
                    net_id = sheet.cell(n, 0).value
                    site_id = sheet.cell(n, 1).value
                    cell_id = sheet.cell(n, 2).value
                    # context = []
                    lac = sheet.cell(n, 3).value

                    ci = sheet.cell(n, 4).value
                    lon = sheet.cell(n, 5).value
                    lat = sheet.cell(n, 6).value
                    site_name = sheet.cell(n, 7).value
                    cell_name = sheet.cell(n, 8).value
                    bsc = sheet.cell(n, 9).value
                    ncc = sheet.cell(n, 10).value
                    bcc = sheet.cell(n, 11).value
                    bisc = sheet.cell(n, 12).value
                    bcch = sheet.cell(n, 16).value

                    tch_list = []
                    for t in range(17, 28):
                        # print(t)

                        #pass

                        tch = sheet.cell(n, t).value

                        #print(n,tch,type(tch),len(tch))
                        if tch =='':
                            pass
                        else:
                            tch_list.append(int(tch))
                    key = str(int(net_id)) + "-" + str(int(site_id)) + "-" + str(int(cell_id))
                    key_cell = str(int(lac)) +"-" + str(int(ci))
                    #print(key)
                    target_cel[key] = [lac, ci, lon, lat, bsc, ncc, bcc, bisc, bcch, site_name, cell_name, tch_list]
                    cel_p[key_cell] = [ncc, bcc,bcch]
                    #print(target_cel[key])

            except Exception as e:
                print(num,e)
                #pass
        #print(target_cel)
        #print(target_cel)
        #print(cel_p)
        return target_cel,cel_p
    else:
        pass

def writefile(filename,dict,c_dict,newfile):
    t = datetime.datetime.now()
    dt = t.strftime('%Y%m%d%H%M%S')
    #打开新文件，改写内容
    path_file =exe_path + "\\" + "copyfile"+ "\\" + filename
    book1 = xlrd.open_workbook(filename)

    GGsmCell = book1.sheet_by_name('GGsmCell')
    rows = GGsmCell.col_values(2)
    #print(len(rows), type(rows))
    new_file = openpyxl.load_workbook(exe_path + "\\" + "copyfile"+ "\\" + newfile)
    #new_file = xlsxwriter.Workbook(exe_path + "\\" + "copyfile"+ "\\" + newfile)
    GGsmCell_new = new_file.get_sheet_by_name('GGsmCell')   #     .get_sheet(2)

    #num_list = [n for n in range(5, 5000)]
    # print(num_list)
    for n in range(5,len(rows)):
        #print(n)
        try:
            cur_netid = GGsmCell.cell(n, 2).value
            cur_siteid = GGsmCell.cell(n, 3).value
            cur_cellid = GGsmCell.cell(n, 4).value
            cur_key = str(cur_netid)+"-" +str(cur_siteid) +"-" + str(cur_cellid)
            #print(cur_key,type(cur_key))
            #print(dict)
            if cur_key in dict.keys():
                #print(n,"写入一次",cur_key,int(dict[cur_key][5]),int(dict[cur_key][6]),int(dict[cur_key][7]))
                #print(GGsmCell_new.cell(n, 7).value)
                GGsmCell_new.cell(row=n+1, column=8, value=int(dict[cur_key][5])) # ncc
                #GGsmCell_new.      #write(n,7,dict[cur_key][5])   #.cell(row =n, column =7).value  = dict[cur_key][5]  #.write(1,3,0)
                #GGsmCell_new.write(n,8,dict[cur_key][6])  #cell(row =n, column =8).value  = dict[cur_key][6]
                #GGsmCell_new.write(n,19,dict[cur_key][8])  #cell(row =n, column =19).value = dict[cur_key][8]
                #GGsmCell_new.write(n,1,"M")                #cell(row =n, column =1).value = "M"
                GGsmCell_new.cell(row=n+1, column=9, value=int(dict[cur_key][6])) #bcc
                GGsmCell_new.cell(row=n+1, column=20, value=int(dict[cur_key][8]))  #bcch
                GGsmCell_new.cell(row=n+1, column=2, value="M")
                #GGsmCell_new['A1'] = '47'
            else:
                GGsmCell_new.cell(row=n+1, column=2, value="P")
                #print("本行忽略！")
        except Exception as e:
            print(e)
            pass

    #################################################


    
    #############GTrx
    GTrx = book1.sheet_by_name('GTrx')
    rows = GTrx.col_values(2)

    GTrx_new = new_file.get_sheet_by_name('GTrx')  #.get_sheet(3)
    #GTrx = workbook.get_sheet_by_name(sheet_names[3])
    #num_list = [n for n in range(4, 10000000)]
        # print(num_list)
    for n in range(5,len(rows)):
        try:
            cur_netid = int(GTrx.cell(n, 2).value)
            cur_siteid = int(GTrx.cell(n, 3).value)
            cur_cellid = int(GTrx.cell(n, 4).value)
            cur_trx = int(GTrx.cell(n, 5).value)-2
            #print(cur_trx)
            #print(cur_netid,cur_siteid,cur_cellid,cur_trx)
            #print(type(cur_netid), type(cur_siteid), type(cur_cellid), type(cur_trx))
            cur_key = str(cur_netid) + "-" + str(cur_siteid) + "-" + str(cur_cellid)
            print(cur_key)
            if cur_key in dict.keys():
                #print("写入一次")
                tchlist = dict[cur_key][11]
                #print("test")
                tt = GTrx.cell(n, 7).value
                #print(tt,type(tt))
                #print(tchlist,cur_key)
                if GTrx.cell(n,7).value == '1':
                    GTrx_new.cell(row=n +1, column=15, value=int(dict[cur_key][8]))
                    GTrx_new.cell(row=n + 1, column=13, value=int(dict[cur_key][6]))
                    GTrx_new.cell(row=n+1, column=2, value="M")
                    #print("写入主频")
                else:
                    #tth =tchlist[0].splist(".")
                    #print(tchlist)
                    GTrx_new.cell(row=n + 1, column=15, value=tchlist[cur_trx])  # b
                    GTrx_new.cell(row=n + 1, column=13, value=int(dict[cur_key][6]))
                    GTrx_new.cell(row=n + 1, column=2, value="M")
                    tchlist.remove(tchlist[cur_trx])
                    print("写入TCH",cur_trx,tchlist[cur_trx])
            else:
                GTrx_new.cell(row=n+1, column=2, value="P")
        except Exception as e:
            print(e)
            pass
    #################################################
    #file_name=filename.split("\\")
    #name =file_name.splist(".")[0]
    #filetype = file_name.splist(".")[1]
    #book2.save(name + dt +"." +filetype)
    #############GExternalGsmCell
    E_cell = book1.sheet_by_name('GExternalGsmCell')
    rows = E_cell.col_values(2)

    E_cell_new = new_file.get_sheet_by_name('GExternalGsmCell')  # .get_sheet(3)
    for n in range(5,len(rows)):
        try:
            cur_lac = E_cell.cell(n, 6).value
            cur_celid = E_cell.cell(n, 7).value
            cur_key = str(cur_lac) + "-" + str(cur_celid)
            #print("curkey",cur_key,type(cur_key))
            if cur_key in c_dict.keys():
                E_cell_new.cell(row=n+1, column=10, value=c_dict[cur_key][2])
                E_cell_new.cell(row=n+1, column=11, value=c_dict[cur_key][0])
                E_cell_new.cell(row=n+1, column=12, value=c_dict[cur_key][1])
                E_cell_new.cell(row=n+1, column=2, value="M")
            else:
                E_cell_new.cell(row=n+1, column=2, value="P")

        except Exception as e:
            print(e)

    #new_file.close()
    new_file.save(filename=exe_path + "\\" + "copyfile"+ "\\" + newfile)

def copy_file(file,path):
    t = datetime.datetime.now()
    dt = t.strftime('%Y%m%d%H%M%S')
    filename = file.split(".")[0]
    typename = file.split(".")[1]
    print(filename)
    shutil.copy(file, path + "\\" + filename +"_" + dt +"."+typename)  # 拷贝文件
    newfilename =filename +"_" + dt +"."+typename
    return  newfilename
if __name__ == '__main__':

    bi ,cel_p=readexecel(bi_path)
    #itemlist = list(bi.keys())
    #print(len(itemlist))
    #for item in itemlist :
    #    print(bi[item])
    if isinstance(bsc_file, list):
        for  item  in bsc_file:
            cdict = deepcopy(bi)
            #print("复制字典")

            #print(item)
            new_file = copy_file(item, exe_path + "\\" + "copyfile")
            writefile(item, cdict, cel_p, new_file)

    else:
        cdict = deepcopy(bi)
        new_file = copy_file(bsc_file,exe_path + "\\" + "copyfile")
        writefile(bsc_file, cdict,cel_p,new_file)
        pass


    #print(bi)
